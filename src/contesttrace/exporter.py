from __future__ import annotations

import csv
import datetime as dt
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .db import SQLiteDB
from .utils import ensure_dir, setup_logger


FRIENDLY_COLUMNS = {
    "id": "ID",
    "title": "通知标题",
    "publish_date": "发布时间",
    "detail_url": "详情链接",
    "organizer": "主办单位",
    "competition_level": "赛事级别",
    "target_audience": "参赛对象",
    "signup_deadline": "报名截止时间",
    "submission_deadline": "作品提交截止时间",
    "competition_time": "比赛时间",
    "requirements": "核心参赛要求",
    "awards": "奖项设置",
    "contact": "联系方式",
    "publisher": "发布人/来源",
    "update_time": "更新时间",
    "full_text": "通知全文",
    "crawled_at": "爬取时间",
    "status": "处理状态",
}


def _auto_filename(prefix: str, ext: str) -> str:
    ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{ts}.{ext}"


def _order_columns(rows: List[Dict[str, Any]]) -> List[str]:
    """统一列顺序：优先核心字段，其余按出现顺序补齐。"""

    if not rows:
        return []
    all_keys: List[str] = []
    for r in rows:
        for k in r.keys():
            if k not in all_keys:
                all_keys.append(k)
    ordered = [c for c in FRIENDLY_COLUMNS.keys() if c in all_keys]
    rest = [c for c in all_keys if c not in ordered]
    return ordered + rest


def _friendly_header(col: str) -> str:
    return FRIENDLY_COLUMNS.get(col, col)


def _autosize_worksheet(ws, max_rows_scan: int = 200) -> None:
    """Excel列宽自适应（扫描前N行，避免极端长文本拖慢）。"""

    try:
        for col_cells in ws.columns:
            # 第一行是表头
            max_len = 0
            for i, cell in enumerate(col_cells, start=1):
                if i > max_rows_scan:
                    break
                v = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(v))
            col_letter = get_column_letter(col_cells[0].column)
            ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 60)
    except Exception:
        return


def export_data(
    db: SQLiteDB,
    exports_dir: Path,
    logger,
    *,
    fmt: str = "xlsx",
    scope: str = "all",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> Optional[Path]:
    """
    一键导出：
    - fmt: xlsx / csv
    - scope: all / date_range / unprocessed
    """

    ensure_dir(exports_dir)

    try:
        if scope == "all":
            rows = db.query_all(limit=999999)
            prefix = "all"
        elif scope == "unprocessed":
            rows = db.query_unprocessed(limit=999999)
            prefix = "unprocessed"
        elif scope == "date_range":
            if not start_date or not end_date:
                logger.error("导出失败：date_range 需要 start_date 与 end_date（YYYY-MM-DD）")
                return None
            rows = db.query_by_date_range(start_date, end_date, limit=999999)
            prefix = f"{start_date}_to_{end_date}"
        else:
            logger.error(f"导出失败：未知 scope={scope}")
            return None

        if not rows:
            logger.warning("导出提示：没有可导出的数据（数据库为空或筛选条件无结果）。")
            return None

        columns = _order_columns(rows)

        if fmt.lower() == "csv":
            out = exports_dir / _auto_filename(prefix, "csv")
            # Windows兼容：utf-8-sig 防止Excel打开乱码
            with out.open("w", encoding="utf-8-sig", newline="") as f:
                writer = csv.writer(f)
                writer.writerow([_friendly_header(c) for c in columns])
                for r in rows:
                    writer.writerow([r.get(c, "") for c in columns])
            logger.info(f"导出成功：{out}")
            return out

        if fmt.lower() == "xlsx":
            out = exports_dir / _auto_filename(prefix, "xlsx")
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "竞赛公告"

                ws.append([_friendly_header(c) for c in columns])
                for r in rows:
                    ws.append([r.get(c, "") for c in columns])

                _autosize_worksheet(ws)
                wb.save(out)
                logger.info(f"导出成功：{out}")
                return out
            except PermissionError:
                logger.error(f"导出失败：文件被占用或无权限写入：{out}")
                return None

        logger.error(f"导出失败：不支持的格式 fmt={fmt}（仅支持 xlsx/csv）")
        return None
    except Exception as e:
        logger.error(f"导出异常：{type(e).__name__}: {e}")
        return None


def main() -> None:
    """命令行入口：默认导出全量Excel。"""

    from .config_center import load_config

    cfg = load_config()
    logger = setup_logger(cfg.paths.logs_dir)
    db = SQLiteDB(cfg.paths.db_path, cfg.database.get("table_name") or "contest_announcements")
    db.init_db(logger)

    export_data(db, cfg.paths.exports_dir, logger, fmt="xlsx", scope="all")


if __name__ == "__main__":
    main()

